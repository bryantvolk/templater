from openpyxl import *
from openpyxl.formatting.rule import *
from openpyxl.styles import *
import re

class Workbook_handler(object):
    """class responsible for opening/closing/writing/copying
        within workbooks"""
    keys = ['S/N:', 'Form:', 'Class:', 'Cat #:', 'Meter Type:', 'FW(MM)',
            'FW(N): ', 'FW(T): ', 'FW(RD): ', 'Test Volts:', 'Test Amps:', 
            'Kt :', 'Hz:', 'Accuracy Class', 'HW Ver:', 'MMA:', 
            'Program:', 'Password:', 'Type: ', 'S/N: ', 'HW: ', 'FW: ',
            'Net ID:', 'Project #:', 'Subject:', 'Proj Engineer:', 'Technician:',
            'Date:']
    def __init__(self,_recieved,_template):
        """opens worksheets and saves as member variables"""
        #input is as recieved
        _recieved = load_workbook(filename = _recieved)
        #output is file with template
        _template = load_workbook(filename = _template)
        self.recieved = _recieved
        self.template = _template

    def copy_template(self, new_name):
        """copies template file as new_name"""
        _template = self.template
        r = re.compile("^[Tt].*")
        template_list = filter(r.match, _template.get_sheet_names())
        source = _template.get_sheet_by_name(template_list[0])
        target = _template.copy_worksheet(source)
        redFill = PatternFill(start_color="E6B7B8", end_color='E6B7B8', fill_type='solid')
        target.conditional_formatting.add('B37:B40', FormulaRule(formula=['ISBLANK(B37)'],
            stopIfTrue=False, fill=redFill)) 
        target.title = new_name

    def map_cells(self, sheet):
        """search through worksheet and form dictionary
            to map cell names(ex S/N: and Form:) to cells that
            need to be filled in"""
        values = dict.fromkeys(self.keys)
        for row in sheet.iter_rows():
            for cell in row:
                cellContent = cell.value
                if not cellContent == None:
                    if cellContent in values:
                        values[cellContent] = cell.offset(row = 0, column = 1)
        return values

    def copy_headers(self, name):
        """copies headers from as reieved to template copy"""
        _recieved = self.recieved
        sheet = _recieved.get_sheet_by_name(name)
        new_sheet = self.template.get_sheet_by_name(name)
        as_recieved_dict = self.map_cells(sheet)
        template_dict = self.map_cells(new_sheet)
        for key, value in as_recieved_dict.iteritems():
            if not value == None:
                if not template_dict[key] == None:
                    print "Writing cell:[ %s ] with value:[ %s ] "%(
                            template_dict[key], value.internal_value)
                    new_sheet[template_dict[key].coordinate] = value.internal_value

    def save_sheet(self,directory):
        """saves sheet to directory specified"""
        self.template.save(filename = directory)
